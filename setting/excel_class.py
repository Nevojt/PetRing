from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment

class ExcelTable:
    def __init__(self):
        self.wb = Workbook()
        self.sheet = self.wb.active

        self.setup_column_dimensions()
        self.set_column_headers()
        

    def setup_column_dimensions(self):
        self.sheet.row_dimensions[1].height = 45
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
        widths = [31, 31, 6, 20, 20, 20, 11, 12, 15, 10, 15, 17, 15, 15, 15, 15, 15, 15, 15, 15, 15, 16, 16, 16, 16, 16, 15, 8]

        for column, width in zip(columns, widths):
            self.sheet.column_dimensions[column].width = width

    def set_column_headers(self):
        headers = ['Index', 'Serie', 'R-PET', 'wersja cennika', 'Baza surowcowa PET', 'Baza surowcowa Rpet', 'Neck', 'Weight(g)', 'Capacity(ml)', 'Color', 'Size pallet', 'Pallet Weight(kg)', 'Bottles per pallet', '[10 000-14 999]', '[15 000-19 999]', '[20 000-24 999]', '[25 000-29 999]', '[30 000-39 999]', '[40 000-49 999]', '[50 000-74 999]', '[75 000-99 999]', '[100 000-149 999]', '[150 000-199 999]', '[200 000-299 999]', '[300 000-499 999]', '[500 000-999 999]', '[1000 000 -]', 'Margin']

        for index, header in enumerate(headers):
            cell = self.sheet.cell(row=1, column=index+1)
            cell.value = header
            
    def add_data(self, row, column, data):
        col_num = column_index_from_string(column)
        for i, value in enumerate(data):
            cell = self.sheet.cell(row=row, column=col_num+i)
            cell.value = value
            

            
    def save(self, file_path):
        self.wb.save(file_path)
