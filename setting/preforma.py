import configparser
from datetime import datetime
import sqlite3
from .interface_v4 import Ui_Preforma
from PyQt5.QtWidgets import QMessageBox
from .dialog import *
from .dialog_update import *
from .excel_class import *
from .table_func import TableFunc
from openpyxl import load_workbook
import postgres_db.models as models
from postgres_db.database import get_db, SessionLocal
from sqlalchemy import update, select
import functools




R_PET_PROCENT = 100 #%
P1 = 66.36 # Пакування Oktabina
P3 = 3.30 # Пакування Kosz
COST_START_SUR = 6.39 # zl



class Preforma(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(Preforma, self).__init__(parent)

        # Основне вікно
        self.ui = Ui_Preforma()
        self.ui.setupUi(self)
        
        # діалогове вікно
        self.dialog = QtWidgets.QDialog()
        self.ua = Ui_Dialog()
        self.ua.setupUi(self.dialog)
        
        # діалогове вікно update Pet, R-Pet 
        self.dialog_update = QtWidgets.QDialog()
        self.update = Ui_Dialog_Update()
        self.update.setupUi(self.dialog_update)

        self.ui.comboBox.currentIndexChanged.connect(self.update_combobox2)
        self.ui.comboBox_2.currentIndexChanged.connect(self.update_combobox3)
        self.ui.comboBox.activated.connect(self.update_r_pet_index)
        # self.ui.comboBox.activated.connect(self.update_label_packing_list)
        
        self.ui.comboBox_2.activated.connect(self.update_r_pet_index)
        self.ui.comboBox_3.activated.connect(self.update_r_pet_index)
        # self.ui.comboBox_3.activated.connect(self.update_label_packing_list)
 
        self.fill_combobox1()
        self.fill_combobox_barwnik()
        self.updateComboBox_5()
        self.updateComboBox_6()
        self.update_tablo()


        # Кнопки головного вікна
        self.ui.pushButton_2.clicked.connect(self.update_table_cost)
        self.ui.pushButton_2.clicked.connect(self.update_data)
        
        self.ui.pushButton.clicked.connect(self.update_dialog)
        self.ui.pushButton_3.clicked.connect(self.open_dialog)
        
        # Кнопки діаовогово вікна
        self.ua.pushButton.clicked.connect(self.button_create_excel)
        self.ua.pushButton_2.clicked.connect(self.button_append_table)
        
        self.update.buttonBox.accepted.connect(self.update_buttons)
        self.update.buttonBox.rejected.connect(self.closed_update_dialog)
        
    def update_table_cost(self):
        # self.update_label_packing_list()
        self.cost_start()
        self.total_cost_raw_material()
        self.upgate_packaging()
        self.update_r_pet_index()
        
        
    
        # Подiя створення файла Excel
    def button_create_excel(self):
        print("Create excel")
        self.create_excel()
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText("Excel created")
        msg.setWindowTitle("Create")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        self.dialog.close()
        
        # Подія додавання таблиці до існуючого файлу
    def button_append_table(self):
        self.append_table()
        print("Append table")
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText("Append table")
        msg.setWindowTitle("Append")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        self.dialog.close()
        
        # Функція відкривання діаловогово вікна
    def open_dialog(self):
        lab = self.ui.label_2.text()
        self.ua.label.setText(lab)
        self.dialog.show()
            

    def fill_combobox1(self):
        numer_values = self.initUIPreforma()
        self.ui.comboBox.addItems(numer_values)

    def update_combobox2(self):
        selected_numer = self.ui.comboBox.currentText()
        gwint_values = self.updateComboBox_2(selected_numer)
        self.ui.comboBox_2.clear()
        self.ui.comboBox_2.addItems(gwint_values)

    def update_combobox3(self):
        selected_numer = self.ui.comboBox.currentText()
        selected_gwint = self.ui.comboBox_2.currentText()
        gramatura_values = self.updateComboBox_3(selected_numer, selected_gwint)
        formatted_data = []
        
        for value in gramatura_values:
            if value.is_integer():  # Перевіряємо, чи значення є цілим числом
                formatted_data.append(str(int(value)))
            else:
                formatted_data.append(str(value))

        self.ui.comboBox_3.clear()
        self.ui.comboBox_3.addItems(formatted_data)
        
    def fill_combobox_barwnik(self):
        barwnik = self.initUIBarwwnik()
        self.ui.comboBox_4.addItems(barwnik)
        self.ui.comboBox_4.activated.connect(self.update_r_pet_index)
        
        
    def initUIPreforma(self):
        db = SessionLocal()
        try:
            forma_values = db.query(models.PreformaDB.numer_form).all()
            return sorted(list(set([value[0] for value in forma_values])))  # Витягуємо значення з кортежів
        finally:
            db.close()
        
    def initUIBarwwnik(self):
        
        db = SessionLocal()
        try:
            forma_values = db.query(models.BarwnikDB.kolor_cecha).all()
            return sorted(list(set([value[0] for value in forma_values])))  # Витягуємо значення з кортежів
        finally:
            db.close()
             
        
    def updateComboBox_2(self, numer):
        db = SessionLocal()
        try:
            gwint_values = db.query(models.PreformaDB.gwint).filter_by(numer_form=numer).all()
            data = sorted(list(set([value[0] for value in gwint_values])))
            return data
        finally:
            db.close()
    
    def updateComboBox_3(self, numer, gwint):
        db = SessionLocal()
        try:
            gramatura_values = db.query(models.PreformaDB.gramatura).filter_by(numer_form=numer, gwint=gwint).all()
            data = sorted(list(set([value[0] for value in gramatura_values])))
            return data
        finally:
            db.close()

        
    # Заповнюємо comboBox_5 значеннями
    def updateComboBox_5(self):
        r_pet_list = ["0", "25", "30", "50", "75", "100"]
        self.ui.comboBox_5.addItems(r_pet_list)
        self.ui.comboBox_5.activated.connect(self.update_r_pet_index)
     
     # Заповнюємо comboBox_6 значеннями   
    def updateComboBox_6(self):
        packing = ["1", "3"]
        self.ui.comboBox_6.addItems(packing)
        self.ui.comboBox_6.activated.connect(self.update_r_pet_index)
  
    # Змінюємо індекс згідно кількості R-Pet
    def update_r_pet_index(self):
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"
        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)
        r_pet_index = self.ui.comboBox_5.currentText()
        index = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}-{self.ui.comboBox_4.currentText()}-{packing}-{self.ui.comboBox_6.currentText()}"
        
   
        if r_pet_index == "0":
            self.ui.label_2.setText(index)    
        elif r_pet_index == "25":
            indexW = index[:1] +"W" + index[1:]
            self.ui.label_2.setText(indexW)
        elif r_pet_index == "30":
            indexV = index[:1] +"V" + index[1:]
            self.ui.label_2.setText(indexV)
        elif r_pet_index == "50":
            indexX = index[:1] +"X" + index[1:]
            self.ui.label_2.setText(indexX)
        elif r_pet_index == "75":
            indexY = index[:1] + "Y" + index[1:]
            self.ui.label_2.setText(indexY)
        elif r_pet_index == "100":
            indexZ = index[:1] + "Z" + index[1:]
            self.ui.label_2.setText(indexZ)
       
    
    def cost_start(self):
        color_box_4 = self.ui.comboBox_4.currentText()
        choice_r_pet = self.ui.comboBox_5.currentText()
        gramatura = float(self.ui.comboBox_3.currentText())
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)

        if color_box_4 == "X":
            if int(choice_r_pet) > 0:
                koszt_urochom = (koszt + (wydajnosc * (gramatura / 1000)) * COST_START_SUR)
                koszt_urochom_r_pet = koszt_urochom + 3 * koszt
                return koszt_urochom, koszt_urochom_r_pet
            else:
                koszt_urochom = (koszt + (wydajnosc * (gramatura / 1000)) * COST_START_SUR)
                koszt_urochom_r_pet = 0
                return koszt_urochom, koszt_urochom_r_pet
        else:
            if int(choice_r_pet) > 0:
                koszt_urochom = (koszt + (wydajnosc * (gramatura / 1000)) * COST_START_SUR) * 3
                koszt_urochom_r_pet = koszt_urochom + 3 * koszt
                return koszt_urochom, koszt_urochom_r_pet
            else:
                koszt_urochom = (koszt + (wydajnosc * (gramatura / 1000)) * COST_START_SUR) * 3
                koszt_urochom_r_pet = 0
                return koszt_urochom, koszt_urochom_r_pet
    
    @functools.lru_cache(maxsize=128)
    def _get_preforma_data(self, index_preforma):
        db = SessionLocal()
        try:
            koszt_obj = db.query(models.PreformaDB.koszt).filter_by(index=index_preforma).first()
            wydajnosc_obj = db.query(models.PreformaDB.wydajnosc_na_godzinu).filter_by(index=index_preforma).first()
            packing_obj = db.query(models.PreformaDB.p1_p3).filter_by(index=index_preforma).first()

            koszt = koszt_obj.koszt
            wydajnosc = wydajnosc_obj.wydajnosc_na_godzinu
            packing = packing_obj.p1_p3
            return koszt, wydajnosc, packing
        finally:
            db.close()

                
    def cena_pet(self):
        kurs = self.view_label_kurs().cena_za_kg   
        pet = self.view_label_pet().cena_za_kg
        
        cena_sur = pet * kurs
        float_value = float(cena_sur) / 1000
        return float_value, cena_sur
    
    def cena_r_pet(self):
        kurs = self.view_label_kurs().cena_za_kg
        r_pet = self.view_label_r_pet().cena_za_kg
        
        cena_sur = r_pet * kurs
        float_value = float(cena_sur) / 1000
        return float_value, cena_sur
    
    # Рахуємо ціну продукції 1000 штук з урахуванням барвника та суровця Pet/R-pet 
    def total_cost_raw_material(self):   # Total Cost Raw material for 1000 pcs
        
        choice_r_pet = int(self.ui.comboBox_5.currentText()) # R-Pet procentent
        kurs_pet = self.cena_pet()[0]
        kurs_r_pet = self.cena_r_pet()[0]

        if choice_r_pet == 0:
            total_result = kurs_pet
        elif choice_r_pet > 0 and choice_r_pet <= 99: 
            fomula_pet = (R_PET_PROCENT - choice_r_pet)
            cost_pet = kurs_pet * (choice_r_pet / 100)
            input_cost = kurs_pet - cost_pet
            cost_pet = kurs_r_pet * (fomula_pet / 100)
            input_cost_rpet = kurs_r_pet - cost_pet
            total_result = input_cost + input_cost_rpet
        elif choice_r_pet == 100:
            total_result = kurs_r_pet
        return total_result
        

    
    # @functools.lru_cache(maxsize=12)
    def total_cost_raw_color(self):
        choice_color = self.ui.comboBox_4.currentText()
        choice_gram = float(self.ui.comboBox_3.currentText())
        total_result = self.total_cost_raw_material()

        cena, dozovanie = self._get_cena_and_dozovanie(choice_color)
        result_material = self._calculate_result_material(cena, dozovanie, choice_gram)

        total_cost_surowiec = choice_gram * total_result
        total_cost_raw = total_cost_surowiec + result_material
        total_cost_tys = round(total_cost_raw, 4)
        
        return total_cost_tys, result_material, total_cost_surowiec

    @functools.lru_cache(maxsize=128)
    def _get_cena_and_dozovanie(self, choice_color):
        db = SessionLocal()
        try:
            cena_obj = db.query(models.BarwnikDB.cena_za_kg).filter_by(kolor_cecha=choice_color).first()
            dozovanie_obj = db.query(models.BarwnikDB.dozowanie).filter_by(kolor_cecha=choice_color).first()
            cena = cena_obj.cena_za_kg
            dozovanie = dozovanie_obj.dozowanie
            return cena, dozovanie
        finally:
            db.close()

    def _calculate_result_material(self, cena, dozovanie, choice_gram):
        if dozovanie > 0:
            result_il_barwnika = choice_gram * (dozovanie / 100)
            result_material = result_il_barwnika * cena / 2
        else:
            result_material = 0
        return result_material



    # Визначаємо ціну за опакованню за 1000 штук
    # @functools.lru_cache(maxsize=128)
    def upgate_packaging(self):
        choice_packaging = self.ui.comboBox_6.currentText()
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)

        if choice_packaging == "1":
            result = P1 * 1000 / packing
        elif choice_packaging == "3":
            result = P3 * 1000 / packing
        result = round(result, 4)
        return result

    
    # Cost Machine for 1000 pcs, розрахунок коштів потрачених на машину за 1000 штук
    # @functools.lru_cache(maxsize=1)
    def cost_machine(self):
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)

        cost_machines = (koszt / wydajnosc) * 1000
        return float(cost_machines)

        
    # Створити функцію, яка оновлює дані для комірки
    def update_data(self):
        # Оновлення даних
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)
        packing_cost = str(round(self.upgate_packaging(), 2))
        cost_machin_tys = str(round(self.cost_machine(), 2))
        cost_color = round(self.total_cost_raw_color()[1], 2)
        many_packaging = str(self.quality_cost())
        quantity_product = str(self.quality_cost_product())
        weight_packaging = str(round(self.total_weight(), 2))
        cost_raw_machines = str(round(self.total_cost_raw_color()[2], 2))
        waste_preform = str(round(self.waste_for_preforms(), 2))
        cost_color_batch = str(round(self.cost_color_batch_waste(), 2))
        total_cost_raw = str(round(self.total_cost_raw_material_tys(), 2))
        cost_start_machine = str(round(self.cost_start_machine(), 2))
        totat_cost_start = str(round(self.total_cost_machine(), 2))
        cost_narzut = str(round(self.total_cost_narzut(), 2))
        finish = str(round(self.total_cost_finish(), 2))
        euro_cost = str(round(self.euro_cost(), 2))
 
            
        
        self.ui.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(str(packing)))
        self.ui.tableWidget.setItem(1, 1, QtWidgets.QTableWidgetItem(many_packaging))
        self.ui.tableWidget.setItem(2, 1, QtWidgets.QTableWidgetItem(quantity_product))
        self.ui.tableWidget.setItem(3, 1, QtWidgets.QTableWidgetItem(weight_packaging))
        self.ui.tableWidget.setItem(4, 1, QtWidgets.QTableWidgetItem(f"{cost_raw_machines} zl"))
        self.ui.tableWidget.setItem(5, 1, QtWidgets.QTableWidgetItem(f"{waste_preform} zl"))
        self.ui.tableWidget.setItem(6, 1, QtWidgets.QTableWidgetItem(f"{cost_color} zl"))
        self.ui.tableWidget.setItem(7, 1, QtWidgets.QTableWidgetItem(f"{cost_color_batch} zl"))
        self.ui.tableWidget.setItem(8, 1, QtWidgets.QTableWidgetItem(f"{total_cost_raw} zl"))
        self.ui.tableWidget.setItem(9, 1, QtWidgets.QTableWidgetItem(f"{packing_cost} zl"))
        self.ui.tableWidget.setItem(10, 1, QtWidgets.QTableWidgetItem(f"{cost_start_machine} zl"))
        self.ui.tableWidget.setItem(11, 1, QtWidgets.QTableWidgetItem(f"{cost_machin_tys} zl"))
        self.ui.tableWidget.setItem(12, 1, QtWidgets.QTableWidgetItem(f"{totat_cost_start} zl"))
        self.ui.tableWidget.setItem(13, 1, QtWidgets.QTableWidgetItem(f"{cost_narzut} zl"))
        self.ui.tableWidget.setItem(14, 1, QtWidgets.QTableWidgetItem(f"{finish} zl"))
        self.ui.tableWidget.setItem(15, 1, QtWidgets.QTableWidgetItem(f"{finish} zl"))
        self.ui.tableWidget.setItem(16, 1, QtWidgets.QTableWidgetItem(f"{euro_cost} €"))

        
        
    # Зчитуємо кількість преформи потрібно облічити
    def quality_cost(self):
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)
        
        try:
            quality = self.ui.lineEdit_2.text()
            
        except (AttributeError, ValueError):
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Critical)
            msg_box.setText("Invalid or missing quality value.\nIlosc dla klienta musi być wpisana\n ilosc Default - 10000")
            msg_box.setWindowTitle("Warning")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec()
            quality = 100000
            many_packaging = quality // packing # Отримуємо калькість опакувань
            return many_packaging
        else:
            many_packaging = int(quality) // packing # Отримуємо калькість опакувань
            return many_packaging
    
    # Облічуємо кількість праформи в опакуванях
    def quality_cost_product(self):
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"
        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)
        
        many_packaging = int(self.quality_cost())
        quantity_product = many_packaging * packing
        return quantity_product
    
    # Вага преформи з опакуваням
    def total_weight(self):
        oktabina = 14
        kosz = 67
        index_preforma = f"{self.ui.comboBox.currentText()}-{self.ui.comboBox_2.currentText()}-{self.ui.comboBox_3.currentText()}"

        koszt, wydajnosc, packing = self._get_preforma_data(index_preforma)
        choice_gram = float(self.ui.comboBox_3.currentText())
        choice_packaging = self.ui.comboBox_6.currentText()
        weight_preform = packing * (choice_gram / 1000)
        
        if choice_packaging == "1":
            weight_packaging = weight_preform + oktabina
        elif choice_packaging == "3":
            weight_packaging = weight_preform + kosz
        return weight_packaging
        
    def waste_for_preforms(self):
        cost_raw_machines = float(self.total_cost_raw_color()[2])
        waste_preform = cost_raw_machines * 3 / 100
        return waste_preform
    
    def cost_color_batch_waste(self):
        cost_color = float(self.total_cost_raw_color()[1])
        cost_color_batch = cost_color * 4.5 / 100
        return cost_color_batch
    

    
    
    def total_cost_raw_material_tys(self):
        cost_color, cost_raw_machines = self.total_cost_raw_color()[1:3]
        waste_preform = self.waste_for_preforms()
        cost_color_batch = self.cost_color_batch_waste()
        
        result = cost_color + cost_raw_machines + waste_preform + cost_color_batch
        return result

    
    def cost_start_machine(self):
        cost_start_pet = self.cost_start()[0]
        cost_start_r_pet = self.cost_start()[1]
        quality = int(self.ui.lineEdit_2.text())

        
        start = max(cost_start_pet, cost_start_r_pet)
        result = start / quality * 1000
        return result
    
    def total_cost_machine(self):
        cost_start = self.cost_start_machine()
        cost_machin = self.cost_machine()
        result = cost_start + cost_machin
        return round(result, 4)
    
    def total_cost_narzut(self):
        narzut = self.view_label_narzut().cena_za_kg
        
        total_cost = round(self.total_cost_machine(), 4)
        result = total_cost + (total_cost * narzut / 100)
        return result 
    
    def total_cost_finish(self):
        total_cost_raw = self.total_cost_raw_material_tys()
        packaging = self.upgate_packaging()
        narzut = self.total_cost_narzut()
        result = total_cost_raw + packaging + narzut
        return round(result, 2)
    
    def euro_cost(self):
        kurs = self.view_label_kurs().cena_za_kg
        total_cost_finish = self.total_cost_finish()
        result = total_cost_finish / kurs
        return result
    
    # Update cost surowiec and kurs, narzut
    def update_dialog(self):
        self.dialog_update.show()
        
    def closed_update_dialog(self):
        self.dialog_update.close()
        
    def update_buttons(self):
        self.date_update()
        edit = self.update.lineEdit.text()
        edit_2 = self.update.lineEdit_2.text()
        edit_3 = self.update.lineEdit_3.text()
        edit_4 = self.update.lineEdit_4.text()
        if len(edit) > 0:
            self.update_kurs()
        if len(edit_2) > 0:
            self.update_pet()
        if len(edit_3) > 0:
            self.update_r_pet()
        if len(edit_4) > 0:
            self.update_narzut()
        else:
            print("No update")
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText("Database update")
        msg.setWindowTitle("Update")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        self.update.lineEdit.clear()
        self.update.lineEdit_2.clear()
        self.update.lineEdit_3.clear()
        self.update.lineEdit_4.clear()
    
    @functools.lru_cache(maxsize=128)  # Застосовуємо кешування до функції
    def view_label_kurs(self, surowiec: str = 'EURO'):
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            return post
        finally:
            db.close()
            
    @functools.lru_cache(maxsize=128)  # Застосовуємо кешування до функції
    def view_label_r_pet(self, surowiec: str = 'R-Pet'):
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            return post
        finally:
            db.close()
            
    @functools.lru_cache(maxsize=128)  # Застосовуємо кешування до функції        
    def view_label_pet(self, surowiec: str = 'Pet'):
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            return post
        finally:
            db.close()
            
    @functools.lru_cache(maxsize=128)  # Застосовуємо кешування до функції        
    def view_label_narzut(self, surowiec: str = 'Narzut'):
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            return post
        finally:
            db.close()
            
            
    @functools.lru_cache(maxsize=128)  # Застосовуємо кешування до функції        
    def date_time(self, surowiec: str = 'Date_Time'):
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            return post
        finally:
            db.close()
            
            
    def update_tablo(self):
        
        # Показуємо актуальну ціну EUR/PLN
        kurs = self.view_label_kurs().cena_za_kg
        self.ui.label_14.setText(str(kurs))
        
        # Показуємо актуальну ціну Pet
        pet = self.view_label_pet().cena_za_kg
        self.ui.label_5.setText(str(pet))
        
        # Показуємо актуальну ціну R-Pet
        r_pet = self.view_label_r_pet().cena_za_kg
        self.ui.label_11.setText(str(r_pet))
        
        # Показуємо процент накладних витрат
        narzut = self.view_label_narzut().cena_za_kg
        self.ui.label_7.setText(str(narzut))
        
        # Показуємо коли були останні зміни
        date = self.date_time().date_time
        self.ui.label_12.setText(str(date))

       
    def update_kurs(self, surowiec: str = 'EURO'):
        cena = float(self.update.lineEdit.text())
    
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            if post:
                post.cena_za_kg = cena  # Змінити поле на нове значення
                db.commit()  # Зберегти зміни у базі даних
            else:
                print("Рядок з таким значенням не знайдений")
        finally:
            db.close()
        
    def update_pet(self, surowiec: str = 'Pet'):
        cena = float(self.update.lineEdit_2.text())
    
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            if post:
                post.cena_za_kg = cena  # Змінити поле на нове значення
                db.commit()  # Зберегти зміни у базі даних
            else:
                print("Рядок з таким значенням не знайдений")
        finally:
            db.close()
        

    # Змінюємо ціну суровца
    def update_r_pet(self, surowiec: str = 'R-Pet'):
        cena = float(self.update.lineEdit_3.text())
    
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            if post:
                post.cena_za_kg = cena  # Змінити поле на нове значення
                db.commit()  # Зберегти зміни у базі даних
            else:
                print("Рядок з таким значенням не знайдений")
        finally:
            db.close()

    # Змінюємо процент накладних витрат
    def update_narzut(self, surowiec: str = 'Narzut'):
        cena = float(self.update.lineEdit_4.text())
    
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            if post:
                post.cena_za_kg = cena  # Змінити поле на нове значення
                db.commit()  # Зберегти зміни у базі даних
            else:
                print("Рядок з таким значенням не знайдений")
        finally:
            db.close()


        
    def date_update(self, surowiec: str = 'Date_Time'):
        date_time = datetime.now()
        # date = date_time.strftime("%d-%m-%y %H:%M")
    
        db = SessionLocal()
        try:
            post = db.query(models.Post).filter(models.Post.surowiec == surowiec).first()
            if post:
                post.date_time = date_time  # Змінити поле на нове значення
                db.commit()  # Зберегти зміни у базі даних
            else:
                print("Рядок з таким значенням не знайдений")
        finally:
            db.close()
        
        
        
        # Блок створення excel файлу
    def create_excel(self):
        r_pet = ('0%', '25%', '30%', '50%', '75%', '100%')
        self.table_func = TableFunc(self)
        # cena_r_pet = self.cena_r_pet()[1]
        
        
        config = configparser.ConfigParser()
        config.read('config.ini')

        last_filename = config.get('File', 'LastFilename')
        parts = last_filename.split("-")
        number = int(parts[1]) + 1
        new_filename = "C:/pdf_files/" + parts[0] + "-" + str(number) + "-" + parts[2]
        second_filename = new_filename[13:]

        
        index = self.table_func.index_one()
        
        # Блок файлу
        table = ExcelTable()
        # Перший та другий стовпчик
        for i, value in enumerate(index):
            table.add_data(i+2, 'A', [value])
            table.add_data(i+2, 'B', [index[0]])
            
        # Третій стовпчик
        for i, value in enumerate(r_pet):
            table.add_data(i+2, 'C', [value])
            
        #  Date time
        date = self.table_func.date_time()
        for i, value in enumerate(date):
            table.add_data(i+2, 'D', [value])
            
        # цінa Pet
        pet = self.table_func.index_E()
        for i, value in enumerate(pet):
            table.add_data(i+2, 'E', [value])
            
        # цінa R-Pet
        r_pets = self.table_func.index_F()
        for i, value in enumerate(r_pets):
            table.add_data(i+2, 'F', [value])
            
        # Neck
        neck = self.table_func.index_G()
        for i, value in enumerate(neck):
            table.add_data(i+2, 'G', [value])
            
        
        # Кількість грам
        gram = self.table_func.index_H()
        for i, value in enumerate(gram):
            table.add_data(i+2, 'H', [value])
            
        # Кількість мілілітрів
        ml = self.table_func.index_I()
        for i, value in enumerate(ml):
            table.add_data(i+2, 'I', [value])
            
        # Колір преформи
        color = self.table_func.index_J()
        for i, value in enumerate(color):
            table.add_data(i+2, 'J', [value])
            
        # Розмір палети
        palet = self.table_func.index_K()
        for i, value in enumerate(palet):
            table.add_data(i+2, 'K', [value])
            
        #  Вага плети(kg)
        weight = self.table_func.index_L()
        for i, value in enumerate(weight):
            table.add_data(i+2, 'L', [value])
            
        # Кількість товару на палеті aбо кощі
        bottles = self.table_func.index_M()
        for i, value in enumerate(bottles):
            table.add_data(i+2, 'M', [value])
          
        # Блок  заповненя ціною  
        list_N = self.table_func.list_N()
        for i, value in enumerate(list_N):
            table.add_data(i+2, 'N', [value])
            
        list_O = self.table_func.list_O()
        for i, value in enumerate(list_O):
            table.add_data(i+2, 'O', [value])
            
        list_P = self.table_func.list_P()
        for i, value in enumerate(list_P):
            table.add_data(i+2, 'P', [value])
     
        list_Q = self.table_func.list_Q()
        for i, value in enumerate(list_Q):
            table.add_data(i+2, 'Q', [value])
        
        list_R = self.table_func.list_R()
        for i, value in enumerate(list_R):
            table.add_data(i+2, 'R', [value])
            
        list_S = self.table_func.list_S()
        for i, value in enumerate(list_S):
            table.add_data(i+2, 'S', [value])
        
        list_T = self.table_func.list_T()
        for i, value in enumerate(list_T):
            table.add_data(i+2, 'T', [value])
            
        list_U = self.table_func.list_U()
        for i, value in enumerate(list_U):
            table.add_data(i+2, 'U', [value])
            
        list_V = self.table_func.list_V()
        for i, value in enumerate(list_V):
            table.add_data(i+2, 'V', [value])
            
        list_W = self.table_func.list_W()
        for i, value in enumerate(list_W):
            table.add_data(i+2, 'W', [value])
        
        list_X = self.table_func.list_X()
        for i, value in enumerate(list_X):
            table.add_data(i+2, 'X', [value])
        
        list_Y = self.table_func.list_Y()
        for i, value in enumerate(list_Y):
            table.add_data(i+2, 'Y', [value])
        
        list_Z = self.table_func.list_Z()
        for i, value in enumerate(list_Z):
            table.add_data(i+2, 'Z', [value])
        
        list_AA = self.table_func.list_AA()
        for i, value in enumerate(list_AA):
            table.add_data(i+2, 'AA', [value])
            
        index_AB = self.table_func.index_AB()
        for i, value in enumerate(index_AB):
            table.add_data(i+2, 'AB', [value])
            
        table.save(second_filename)
        
        
        # Збереження нового імені файлу
        config.set('File', 'LastFilename', second_filename)
        with open('config.ini', 'w') as config_file:
            config.write(config_file)
        
    def append_table(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        last_filename = config.get('File', 'LastFilename')
            
        self.table_func = TableFunc(self)
        index = self.table_func.index_one()
        # Завантажуємо існуючу книгу
        wb = load_workbook(filename = last_filename)

        # Вибираємо активний лист (або ви можете вибрати лист за назвою wb.get_sheet_by_name('Sheet1'))
        ws = wb.active 

        start_row = ws.max_row + 1
        
        next_row = start_row
        for value in index:
            ws.cell(row=next_row, column=1, value=value)
            next_row += 1
            
        next_row = start_row  
        for value in index:
            ws.cell(row=next_row, column=2, value=index[0])
            next_row += 1
            
        # Третій стовпчик
        next_row = start_row
        r_pet = ('0%', '25%', '30%', '50%', '75%', '100%')
        for value in r_pet:
            ws.cell(row=next_row, column=3, value=value)
            next_row += 1

            
        #  Date time
        next_row = start_row
        date = self.table_func.date_time()
        for value in date:
            ws.cell(row=next_row, column=4, value=value)
            next_row += 1
            
        # цінa Pet
        next_row = start_row
        pet = self.table_func.index_E()
        for value in pet:
            ws.cell(row=next_row, column=5, value=value)
            next_row += 1
            
        # цінa R-Pet
        next_row = start_row
        r_pets = self.table_func.index_F()
        for value in r_pets:
            ws.cell(row=next_row, column=6, value=value)
            next_row += 1
            
        # Neck
        next_row = start_row
        neck = self.table_func.index_G()
        for value in neck:
            ws.cell(row=next_row, column=7, value=value)
            next_row += 1
            
        
        # Кількість грам
        next_row = start_row
        gram = self.table_func.index_H()
        for value in gram:
            ws.cell(row=next_row, column=8, value=value)
            next_row += 1
            
        # Кількість мілілітрів
        next_row = start_row
        ml = self.table_func.index_I()
        for value in ml:
            ws.cell(row=next_row, column=9, value=value)
            next_row += 1
            
        # Колір преформи
        next_row = start_row
        color = self.table_func.index_J()
        for value in color:
            ws.cell(row=next_row, column=10, value=value)
            next_row += 1
            
        # Розмір палети
        next_row = start_row
        palet = self.table_func.index_K()
        for value in palet:
            ws.cell(row=next_row, column=11, value=value)
            next_row += 1
            
        #  Вага плети(kg)
        next_row = start_row
        weight = self.table_func.index_L()
        for value in weight:
            ws.cell(row=next_row, column=12, value=value)
            next_row += 1
            
        # Кількість товару на палеті aбо кощі
        next_row = start_row
        bottles = self.table_func.index_M()
        for value in bottles:
            ws.cell(row=next_row, column=13, value=value)
            next_row += 1
          
        #  Блок  заповненя ціною  
        next_row = start_row
        list_N = self.table_func.list_N()
        for value in list_N:
            ws.cell(row=next_row, column=14, value=value)
            next_row += 1
            
        next_row = start_row
        list_O = self.table_func.list_O()
        for value in list_O:
            ws.cell(row=next_row, column=15, value=value)
            next_row += 1
        
        next_row = start_row
        list_P = self.table_func.list_P()
        for value in list_P:
            ws.cell(row=next_row, column=16, value=value)
            next_row += 1
            
        next_row = start_row
        list_Q = self.table_func.list_Q()
        for value in list_Q:
            ws.cell(row=next_row, column=17, value=value)
            next_row += 1
            
        next_row = start_row
        list_R = self.table_func.list_R()
        for value in list_R:
            ws.cell(row=next_row, column=18, value=value)
            next_row += 1
        
        next_row = start_row
        list_S = self.table_func.list_S()
        for value in list_S:
            ws.cell(row=next_row, column=19, value=value)
            next_row += 1
        
        next_row = start_row
        list_T = self.table_func.list_T()
        for value in list_T:
            ws.cell(row=next_row, column=20, value=value)
            next_row += 1
            
        next_row = start_row
        list_U = self.table_func.list_U()
        for value in list_U:
            ws.cell(row=next_row, column=21, value=value)
            next_row += 1
        
        next_row = start_row
        list_V = self.table_func.list_V()
        for value in list_V:
            ws.cell(row=next_row, column=22, value=value)
            next_row += 1
        
        next_row = start_row
        list_W = self.table_func.list_W()
        for value in list_W:
            ws.cell(row=next_row, column=23, value=value)
            next_row += 1
        
        next_row = start_row
        list_X = self.table_func.list_X()
        for value in list_X:
            ws.cell(row=next_row, column=24, value=value)
            next_row += 1
        
        next_row = start_row
        list_Y = self.table_func.list_Y()
        for value in list_Y:
            ws.cell(row=next_row, column=25, value=value)
            next_row += 1
        
        next_row = start_row
        list_Z = self.table_func.list_Z()
        for value in list_Z:
            ws.cell(row=next_row, column=26, value=value)
            next_row += 1
        
        next_row = start_row
        list_AA = self.table_func.list_AA()
        for value in list_AA:
            ws.cell(row=next_row, column=27, value=value)
            next_row += 1
            
        next_row = start_row
        index_AB = self.table_func.index_AB()
        for value in index_AB:
            ws.cell(row=next_row, column=28, value=value)
            next_row += 1
            
            
        

        # Зберігаємо книгу
        wb.save(filename = last_filename)

        

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ui = Preforma()
    ui.show()
    sys.exit(app.exec_())

    